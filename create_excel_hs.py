"""
Genera el Excel maestro de prueba para Hub Clínico HS Canarias v1.0
Requiere: pip install openpyxl
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

OUTPUT_DIR = r"C:\Users\b32ma\Documents\Canarias\HUC\Hub-Clinico-HS-Canarias"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "Hub_Clinico_HS_Canarias.xlsx")

# 195 cabeceras HS_EXPORT_HEADERS
HS_HEADERS = [
    # Bloque A — Identificación (9)
    'NHC', 'Fecha_Visita', 'Tipo_Visita', 'Centro', 'Consulta', 'Profesional', 'Origen_Paciente', 'CIE10', 'Motivo_Consulta',
    # Bloque B — Datos basales (17)
    'Sexo', 'Edad', 'Anio_Nacimiento', 'Anio_Inicio_Sintomas', 'Anio_Diagnostico', 'Retraso_Diagnostico_Anios', 'Antecedentes_Familiares_HS', 'Fumador', 'Cigarros_Dia', 'Anios_Fumador', 'Exfumador', 'Peso', 'Talla', 'IMC',
    # Bloque C — Comorbilidades (20)
    'Comorbilidad_Diabetes', 'Comorbilidad_HTA', 'Comorbilidad_Dislipemia', 'Comorbilidad_Obesidad', 'Comorbilidad_Sindrome_Metabolico', 'Comorbilidad_ECV', 'Comorbilidad_Esteatosis_Hepatica', 'Comorbilidad_EII', 'Comorbilidad_Crohn', 'Comorbilidad_Colitis_Ulcerosa', 'Comorbilidad_Artritis', 'Comorbilidad_Psoriasis', 'Comorbilidad_Acne', 'Comorbilidad_Sinus_Pilonidal', 'Comorbilidad_SOP', 'Comorbilidad_Depresion', 'Comorbilidad_Ansiedad', 'Comorbilidad_PASH_PAPASH', 'Comorbilidad_Pioderma_Gangrenoso', 'Comorbilidad_Foliculitis_Decalvante', 'Comorbilidad_Otras',
    # Bloque D — Actividad clínica (13)
    'Hurley', 'Nodulos_Total', 'Abscesos_Total', 'Fistulas_Total', 'Fistulas_Drenantes_Total', 'IHS4_Clinico', 'IHS4_Clinico_Categoria', 'Dolor_EVA', 'Prurito_EVA', 'Supuracion_EVA', 'Brotes_Ultimos_3_Meses', 'Visitas_Urgencias_HS_Ultimos_6_Meses', 'Antibioticos_Ultimos_6_Meses',
    # Bloque E — Lesiones por región (49)
    'Axila_Der_Nodulos', 'Axila_Der_Abscesos', 'Axila_Der_Fistulas', 'Axila_Der_Fistulas_Drenantes',
    'Axila_Izq_Nodulos', 'Axila_Izq_Abscesos', 'Axila_Izq_Fistulas', 'Axila_Izq_Fistulas_Drenantes',
    'Inframamaria_Der_Nodulos', 'Inframamaria_Der_Abscesos', 'Inframamaria_Der_Fistulas', 'Inframamaria_Der_Fistulas_Drenantes',
    'Inframamaria_Izq_Nodulos', 'Inframamaria_Izq_Abscesos', 'Inframamaria_Izq_Fistulas', 'Inframamaria_Izq_Fistulas_Drenantes',
    'Inguinal_Der_Nodulos', 'Inguinal_Der_Abscesos', 'Inguinal_Der_Fistulas', 'Inguinal_Der_Fistulas_Drenantes',
    'Inguinal_Izq_Nodulos', 'Inguinal_Izq_Abscesos', 'Inguinal_Izq_Fistulas', 'Inguinal_Izq_Fistulas_Drenantes',
    'Genital_Nodulos', 'Genital_Abscesos', 'Genital_Fistulas', 'Genital_Fistulas_Drenantes',
    'Perineal_Nodulos', 'Perineal_Abscesos', 'Perineal_Fistulas', 'Perineal_Fistulas_Drenantes',
    'Perianal_Nodulos', 'Perianal_Abscesos', 'Perianal_Fistulas', 'Perianal_Fistulas_Drenantes',
    'Glutea_Der_Nodulos', 'Glutea_Der_Abscesos', 'Glutea_Der_Fistulas', 'Glutea_Der_Fistulas_Drenantes',
    'Glutea_Izq_Nodulos', 'Glutea_Izq_Abscesos', 'Glutea_Izq_Fistulas', 'Glutea_Izq_Fistulas_Drenantes',
    'Otra_Region_Nodulos', 'Otra_Region_Abscesos', 'Otra_Region_Fistulas', 'Otra_Region_Fistulas_Drenantes',
    'Otra_Region_Descripcion',
    # Bloque F — Cicatrices (4)
    'Cicatrices', 'Cicatrices_Descripcion', 'Tuneles_Cronicos', 'Limitacion_Funcional', 'Limitacion_Funcional_Descripcion',
    # Bloque G — Ecografía (9)
    'Ecografia_Realizada', 'Eco_Nodulos_Total', 'Eco_Abscesos_Total', 'Eco_Fistulas_Total', 'Eco_Fistulas_Drenantes_Total', 'IHS4_Ecografico', 'IHS4_Ecografico_Categoria', 'Doppler_Positivo', 'Hallazgos_Ecograficos',
    # Bloque H — PROMs (10)
    'DLQI', 'HSQoL24', 'EVA_Dolor_Paciente', 'EVA_Impacto_Global', 'EVA_Olor', 'EVA_Supuracion', 'Dias_Baja_Ultimos_6_Meses', 'Impacto_Laboral', 'Impacto_Sexual', 'Comentarios_PROMs',
    # Bloque I — Tratamiento (24)
    'Tratamiento_Actual', 'Fecha_Inicio_Tratamiento_Actual', 'Previo_Antibiotico_1', 'Previo_Antibiotico_Dosis_1', 'Previo_Antibiotico_2', 'Previo_Antibiotico_Dosis_2', 'Previo_Antibiotico_3', 'Previo_Antibiotico_Dosis_3', 'Previo_Biologico_1', 'Previo_Biologico_Dosis_1', 'Previo_Biologico_2', 'Previo_Biologico_Dosis_2', 'Previo_Biologico_3', 'Previo_Biologico_Dosis_3', 'Previo_Otro_Sistemico_1', 'Previo_Otro_Sistemico_Dosis_1', 'Previo_Otro_Sistemico_2', 'Previo_Otro_Sistemico_Dosis_2', 'Trat_Antibiotico', 'Trat_Antibiotico_Dosis', 'Trat_Biologico', 'Trat_Biologico_Dosis', 'Trat_Topico', 'Trat_Topico_Dosis', 'Trat_Otro', 'Trat_Otro_Dosis',
    # Bloque J — Decisión terapéutica (10)
    'Decision_Terapeutica', 'Continuar_Tratamiento', 'Optimizar_Adherencia', 'Ajuste_Dosis', 'Cambio_Tratamiento', 'Cambio_Motivo', 'Efectos_Adversos', 'Efectos_Adversos_Descripcion', 'Suspension_Tratamiento', 'Suspension_Motivo',
    # Bloque K — Cirugía/comité (14)
    'Requiere_Cirugia', 'Derivacion_Dermatologia_Quirurgica', 'Derivacion_Cirugia_General', 'Derivacion_Cirugia_Plastica', 'Derivacion_Ginecologia', 'Derivacion_Urologia', 'Region_Quirurgica', 'Prioridad_Quirurgica', 'Cirugia_Realizada', 'Fecha_Cirugia', 'Resultado_Cirugia', 'Comite_Multidisciplinar', 'Decision_Comite', 'Comentarios_Cirugia_Comite',
    # Bloque L — Analítica/preventiva (9)
    'Analitica_Solicitada', 'Analitica_Realizada', 'Medicina_Preventiva_Solicitada', 'Medicina_Preventiva_Realizada', 'Vacunacion_Revisada', 'Cribado_TB', 'Cribado_Hepatitis', 'Cribado_VIH', 'Observaciones_Preventiva',
    # Bloque M — Cierre (6)
    'Plan_Clinico', 'Fecha_Proxima_Revision', 'Comentarios_Adicionales', 'TXT_Historia_Generado', 'Fecha_Exportacion', 'Version_App'
]

def style_header_row(ws, row=1):
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="008777", end_color="008777", fill_type="solid")
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def create_workbook():
    wb = openpyxl.Workbook()

    # ─── Hoja 1: HS (195 cabeceras) ───
    ws_hs = wb.active
    ws_hs.title = "HS"
    for col_idx, header in enumerate(HS_HEADERS, 1):
        ws_hs.cell(row=1, column=col_idx, value=header)
    style_header_row(ws_hs)
    print(f"  [OK] Hoja 'HS': {len(HS_HEADERS)} cabeceras")

    # ─── Hoja 2: Profesionales ───
    ws_prof = wb.create_sheet("Profesionales")
    ws_prof.cell(1, 1, "Nombre_Completo")
    ws_prof.cell(1, 2, "Cargo")
    ws_prof.cell(1, 3, "Activo")
    style_header_row(ws_prof)
    profesionales = [
        ("Vania Lukoviek", "Dermatología", "SI"),
        ("Marta García", "Dermatología", "SI"),
        ("Kiko Guimerá", "Dermatología", "SI"),
    ]
    for i, (nombre, cargo, activo) in enumerate(profesionales, 2):
        ws_prof.cell(i, 1, nombre)
        ws_prof.cell(i, 2, cargo)
        ws_prof.cell(i, 3, activo)
    print(f"  [OK] Hoja 'Profesionales': {len(profesionales)} registros")

    # ─── Hoja 3: Consultas ───
    ws_cons = wb.create_sheet("Consultas")
    ws_cons.cell(1, 1, "Consulta")
    ws_cons.cell(1, 2, "Activo")
    style_header_row(ws_cons)
    consultas = [
        ("Consulta monográfica HS", "SI"),
        ("Dermatología general", "SI"),
        ("Consulta multidisciplinar", "SI"),
    ]
    for i, (consulta, activo) in enumerate(consultas, 2):
        ws_cons.cell(i, 1, consulta)
        ws_cons.cell(i, 2, activo)
    print(f"  [OK] Hoja 'Consultas': {len(consultas)} registros")

    # ─── Hoja 4: Farmacos_HS ───
    ws_farm = wb.create_sheet("Farmacos_HS")
    ws_farm.cell(1, 1, "Farmaco")
    ws_farm.cell(1, 2, "Categoria")
    ws_farm.cell(1, 3, "Principio_Activo")
    ws_farm.cell(1, 4, "Activo")
    style_header_row(ws_farm)
    farmacos = [
        # Antibióticos
        ("Clindamicina + Rifampicina", "Antibiotico", "Clindamicina/Rifampicina", "SI"),
        ("Doxiciclina", "Antibiotico", "Doxiciclina", "SI"),
        ("Minociclina", "Antibiotico", "Minociclina", "SI"),
        ("Eritromicina", "Antibiotico", "Eritromicina", "SI"),
        # Biológicos
        ("Adalimumab", "Biologico", "Adalimumab", "SI"),
        ("Infliximab", "Biologico", "Infliximab", "SI"),
        ("Secukinumab", "Biologico", "Secukinumab", "SI"),
        ("Bimekizumab", "Biologico", "Bimekizumab", "SI"),
        # Tópicos
        ("Clindamicina tópica", "Topico", "Clindamicina", "SI"),
        ("Resorcina", "Topico", "Resorcina", "SI"),
        ("Clorhexidina", "Topico", "Clorhexidina", "SI"),
        # Sistémicos No Biológicos
        ("Acitretina", "Sistemico_No_Biologico", "Acitretina", "SI"),
        ("Metotrexato", "Sistemico_No_Biologico", "Metotrexato", "SI"),
        ("Ciclosporina", "Sistemico_No_Biologico", "Ciclosporina", "SI"),
        ("Apremilast", "Sistemico_No_Biologico", "Apremilast", "SI"),
        # Otros
        ("Espironolactona", "Otros", "Espironolactona", "SI"),
        ("Finasterida", "Otros", "Finasterida", "SI"),
        ("Metformina", "Otros", "Metformina", "SI"),
    ]
    for i, (farmaco, categoria, pa, activo) in enumerate(farmacos, 2):
        ws_farm.cell(i, 1, farmaco)
        ws_farm.cell(i, 2, categoria)
        ws_farm.cell(i, 3, pa)
        ws_farm.cell(i, 4, activo)
    print(f"  [OK] Hoja 'Farmacos_HS': {len(farmacos)} farmacos")

    # ─── Hoja 5: Instrucciones ───
    ws_inst = wb.create_sheet("Instrucciones")
    instrucciones = [
        ("Hoja", "Descripción"),
        ("HS", "Datos clínicos de pacientes con Hidradenitis Suppurativa. 195 columnas. Una fila por visita."),
        ("Profesionales", "Listado de profesionales. Columnas: Nombre_Completo, Cargo, Activo (SI/NO)."),
        ("Consultas", "Listado de consultas. Columnas: Consulta, Activo (SI/NO)."),
        ("Farmacos_HS", "Catálogo de fármacos HS. Columnas: Farmaco, Categoria, Principio_Activo, Activo."),
        ("", ""),
        ("Uso", ""),
        ("1.", "Cargar este Excel en index.html del Hub Clínico HS Canarias."),
        ("2.", "Seleccionar profesional en el session gate."),
        ("3.", "Usar primera_visita.html para registrar nueva visita HS."),
        ("4.", "Usar seguimiento.html para visitas de seguimiento."),
        ("5.", "Generar TXT para historia clínica → luego copiar fila TSV para la hoja HS."),
        ("6.", "Pegar la fila TSV en la hoja HS del archivo maestro."),
        ("", ""),
        ("Categorías de fármacos", ""),
        ("Antibiotico", "Antibióticos sistémicos (clindamicina+rifampicina, doxiciclina, etc.)"),
        ("Biologico", "Biológicos (adalimumab, infliximab, secukinumab, etc.)"),
        ("Topico", "Tratamientos tópicos (clindamicina tópica, resorcina, etc.)"),
        ("Sistemico_No_Biologico", "Sistémicos no biológicos (acitretina, metotrexato, etc.)"),
        ("Otros", "Otros tratamientos (espironolactona, finasterida, metformina, etc.)"),
    ]
    for i, (col1, col2) in enumerate(instrucciones, 1):
        ws_inst.cell(i, 1, col1)
        ws_inst.cell(i, 2, col2)
        if i == 1:
            ws_inst.cell(i, 1).font = Font(bold=True)
            ws_inst.cell(i, 2).font = Font(bold=True)
    ws_inst.column_dimensions['A'].width = 28
    ws_inst.column_dimensions['B'].width = 80
    print("  [OK] Hoja 'Instrucciones'")

    # Guardar
    wb.save(OUTPUT_FILE)
    print(f"\n[OK] Excel guardado: {OUTPUT_FILE}")

if __name__ == "__main__":
    print("Generando Excel maestro HS Canarias...")
    create_workbook()
